#!/usr/bin/env python3
"""สร้างไฟล์ ImportData.gs จากไฟล์สำรอง .json ของแอปเดิม

    python3 tools/make_import.py <ไฟล์สำรอง.json> <sheet.xlsx> [ไฟล์ผลลัพธ์]

ไฟล์ผลลัพธ์มีชื่อ เบอร์ และที่อยู่ลูกค้าจริง — **ห้าม commit ขึ้น repo เด็ดขาด**
ส่งให้เจ้าของร้านตรง ๆ ให้วางลงโปรเจกต์ Apps Script แล้วลบทิ้งหลังนำเข้าเสร็จ

หลักที่ยึด
  - **ชีทเป็นความจริง** รหัส SKU ของเคมีในแอปเดิมชี้ไปคนละตัวกับในชีท 5 จาก 14 ล็อต
    จึงจับคู่ด้วย "ชนิดสินค้า + ขนาดบรรจุ" ไม่ใช่ด้วยรหัส
  - เอาเฉพาะออเดอร์ของวันล่าสุดวันเดียว ที่เหลือไม่นำเข้า
  - ทุกอย่างที่เดาไม่ได้ ให้ล้มทันทีพร้อมบอกว่าติดตรงไหน ไม่เดาแล้วเขียนลงชีท
"""
import json
import re
import sys
import pathlib

# ---------------------------------------------------------------- จับคู่สินค้า

def norm(s):
    s = str(s or '').lower()
    s = s.replace('×', '*').replace('✕', '*').replace('x', '*')
    s = re.sub(r'[()\[\]]', ' ', s)
    s = re.sub(r'[-–—/,]', '*', s)
    s = re.sub(r'\s+', '', s)
    return re.sub(r'\*+', '*', s)


def digits(s):
    return re.findall(r'\d+(?:\.\d+)?', str(s or '').replace(',', ''))


def chem_kind(s):
    """ชนิดน้ำยา — ใช้คู่กับขนาดบรรจุแทนรหัส SKU ที่เชื่อไม่ได้"""
    t = str(s).lower()
    for key, tag in (('o-014', 'O-014'), ('acetone', 'Acetone'), ('s-011', 'S-011'),
                     ('pcb', 'PCB'), ('ethanol', 'Ethanol')):
        if key in t:
            return tag
    if 'ipa' in t or 'isopropyl' in t:
        return 'IPA'
    return None


def volume_ml(s):
    t = str(s).lower().replace(',', '')
    m = re.search(r'(\d+(?:\.\d+)?)\s*(?:ลิตร|l\b)', t)
    if m:
        return int(float(m.group(1)) * 1000)
    m = re.search(r'(\d+)\s*ml', t)
    return int(m.group(1)) if m else None


def new_sku(name, sheet, have):
    """ออกรหัสใหม่ต่อจากเลขสูงสุดที่มีอยู่ ให้อยู่ในระบบเลขเดียวกับของเดิม"""
    mx = 0
    for s in sheet:
        m = re.match(r'^SKU-(\d+)$', s['sku'])
        if m:
            mx = max(mx, int(m.group(1)))
    n = mx + 1
    while ('SKU-%d' % n) in have:
        n += 1
    return 'SKU-%d' % n


def match_by_name(name, sheet):
    """จับคู่สินค้าทั่วไปด้วยชื่อ — ต้องเหลือตัวเดียวเท่านั้นถึงจะยอมรับ"""
    n = norm(name)
    hit = [s for s in sheet if norm(s['name']) == n]
    if len(hit) == 1:
        return hit[0]
    dn = digits(name)
    if dn:
        cand = [s for s in sheet if digits(s['name']) == dn]
        if len(cand) == 1:
            return cand[0]
        near = [s for s in cand if norm(s['name'])[:6] == n[:6]]
        if len(near) == 1:
            return near[0]
    return None


def match_chem(name, sheet):
    k, v = chem_kind(name), volume_ml(name)
    if not k or not v:
        return None
    hit = [s for s in sheet if chem_kind(s['name']) == k and volume_ml(s['name']) == v]
    return hit[0] if len(hit) == 1 else None


# ------------------------------------------------------------ แปลงค่าให้เข้าชีท

CARRIER = {
    'flash express': 'Flash Express',
    'kerry express': 'Kerry Express',
    'ไปรษณีย์ไทย': 'ไปรษณีย์ไทย',
    'ไปรษณีไทย': 'ไปรษณีย์ไทย',
    'ส่งด่วน': 'ส่งด่วน (ไรเดอร์)',
    'ส่งด่วน (ไรเดอร์)': 'ส่งด่วน (ไรเดอร์)',
    'รับเองที่ร้าน': 'รับเองที่ร้าน',
}


def carrier_of(v):
    return CARRIER.get(str(v or '').strip().lower(), str(v or '').strip())


def status_of(o):
    """สถานะของชีทมาจากสองช่องของแอปเดิมรวมกัน"""
    if o.get('ostatus') == 'cancel':
        return 'ยกเลิก'
    if o.get('ostatus') == 'shipped':
        return 'ส่งแล้ว'
    if o.get('ostatus') == 'done':
        return 'ส่งแล้ว'
    if o.get('status') == 'paid':
        return 'ชำระแล้ว'
    return 'รอชำระ'


def num(v, d=0):
    try:
        return float(str(v).replace(',', ''))
    except (TypeError, ValueError):
        return d


def gs(v):
    return json.dumps(v, ensure_ascii=False)


# ------------------------------------------------------------------------ main

def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    backup = pathlib.Path(sys.argv[1])
    xlsx = pathlib.Path(sys.argv[2])
    out = pathlib.Path(sys.argv[3]) if len(sys.argv) > 3 else pathlib.Path('out/ImportData.gs')
    out.parent.mkdir(parents=True, exist_ok=True)

    # ต้นทุน/ราคาขายที่เจ้าของร้านสั่งแก้ อยู่ในไฟล์แยก ไม่ใส่ในสคริปต์
    # เพราะ repo เปิดสาธารณะ และต้นทุนเป็นข้อมูลของร้าน
    # รูปแบบ: [{"sku":"SKU-202","cost":78,"price":225,"why":"..."}]
    prices = []
    if len(sys.argv) > 4:
        prices = json.loads(pathlib.Path(sys.argv[4]).read_text(encoding='utf-8'))

    # คำตัดสินว่าชื่อสินค้าในแอปเดิมตรงกับ SKU ไหนในชีท สำหรับตัวที่จับคู่อัตโนมัติไม่ได้
    # ค่า "NEW" = เป็นสินค้าใหม่จริง ให้เพิ่มเข้าฐานสินค้า
    override = {}
    if len(sys.argv) > 5:
        override = json.loads(pathlib.Path(sys.argv[5]).read_text(encoding='utf-8'))

    # ชีทเวอร์ชันก่อน ใช้เป็น "สะพาน" อย่างเดียว
    # บางบรรทัดในออเดอร์เก็บชื่อไว้แบบตัดหัวทิ้ง ("3.175-22-3.175-45L-1pcs")
    # แต่มีรหัสของตอนนั้นติดมาด้วย จึงเอารหัสไปเปิดชีทเก่าเพื่อหาชื่อเต็ม
    # แล้วค่อยเอาชื่อเต็มไปจับคู่กับชีทใหม่ — ไม่ได้เอารหัสเก่าไปลงชีทโดยตรง
    import openpyxl
    old_name = {}
    if len(sys.argv) > 6:
        wb_old = openpyxl.load_workbook(pathlib.Path(sys.argv[6]), data_only=True)['ฐานสินค้า']
        for r in range(6, 200):
            s = wb_old.cell(r, 2).value
            if s:
                old_name[str(s).strip()] = str(wb_old.cell(r, 4).value)

    import openpyxl
    d = json.loads(backup.read_text(encoding='utf-8'))
    ws = openpyxl.load_workbook(xlsx)['ฐานสินค้า']
    sheet = []
    for r in range(6, 200):
        sku = ws.cell(r, 2).value
        if sku:
            sheet.append({'sku': str(sku).strip(), 'name': str(ws.cell(r, 4).value),
                          'cost': ws.cell(r, 7).value, 'price': ws.cell(r, 8).value})
    have = {s['sku'] for s in sheet}
    P = {p['id']: p for p in d['products']}
    by_old_sku = {}
    for p in d['products']:
        k = str(p.get('sku') or '').strip()
        if k and k not in by_old_sku:
            by_old_sku[k] = p
    problems = []

    # ---- ล็อตเคมี + รับเข้าคู่กัน ----
    lots, recv = [], []
    for l in d.get('lots', []):
        p = P.get(l.get('pid'))
        if not p:
            problems.append('ล็อต %s: ไม่พบสินค้าในไฟล์สำรอง' % l.get('lot'))
            continue
        m = match_chem(p['name'], sheet)
        if not m:
            problems.append('ล็อต %s (%s): จับคู่กับสินค้าในชีทไม่ได้' % (l.get('lot'), p['name']))
            continue
        qty = num(l.get('qty'))
        lots.append({'sku': m['sku'], 'lotNo': l['lot'], 'exp': l.get('exp') or '',
                     'recv': l.get('recv') or '', 'qty': qty,
                     'note': 'ยกมาจากระบบเดิม (รหัสเดิม %s)' % p.get('sku', '-')})
        recv.append({'date': l.get('recv') or '', 'doc': 'MIG-' + l['lot'], 'type': 'ซื้อเข้า',
                     'ref': 'ยกยอดจากระบบเดิม', 'sku': m['sku'], 'qty': qty,
                     'cost': num(l.get('cost')), 'staff': 'ย้ายระบบ',
                     'note': 'ล็อต ' + l['lot']})

    # ---- ออเดอร์: เอาเฉพาะวันล่าสุดวันเดียว ----
    orders_all = [o for o in d.get('orders', []) if o.get('date')]
    if not orders_all:
        sys.exit('ไม่มีออเดอร์ในไฟล์สำรอง')
    today = max(o['date'] for o in orders_all)
    todays = [o for o in orders_all if o['date'] == today]

    orders, new_products = [], []
    for o in todays:
        its = o.get('items') or []
        if not its and o.get('pid'):
            its = [{'pid': o['pid'], 'name': o.get('pname'), 'qty': o.get('qty'),
                    'price': '', 'total': o.get('amt')}]
        if not its:
            problems.append('ออเดอร์ %s: ไม่มีข้อมูลสินค้าเลย' % o.get('no'))
            continue

        lines, subtotal = [], 0.0
        for ix, it in enumerate(its, 1):
            p = P.get(it.get('pid') or '')
            if not p and it.get('sku'):
                # บางบรรทัดเก็บชื่อไว้แบบย่อ ("3.175-22-3.175-45L-1pcs")
                # หารายการเต็มจากคลังสินค้าของแอปเดิมด้วยรหัสของ "แอปเดิม" ก่อน
                # (ใช้แค่หาชื่อเต็ม ไม่ได้เอารหัสนั้นไปลงชีท)
                p = by_old_sku.get(str(it['sku']).strip())
            name = (p or {}).get('name') or it.get('name') or ''

            # ห้ามเชื่อรหัส SKU ที่ติดมากับแอปเดิมเด็ดขาด
            # เจ้าของร้านเรียงรหัสในชีทใหม่ 44 จาก 97 รหัสที่ชื่อเดิมมี ตอนนี้ชี้ไปคนละสินค้า
            # เช่น SKU-172 เคยเป็น Straight Endmill 2F 2.0-17 ตอนนี้เป็น Micro Square Endmill 0.6-1.2
            # จับคู่ด้วยชื่อสินค้าอย่างเดียว ชีทคือความจริง
            m = match_by_name(name, sheet)
            if not m:
                # ชื่อบนบรรทัดถูกตัดหัว — เอารหัสของตอนนั้นไปเปิดชีทเก่าหาชื่อเต็มก่อน
                oldsku = str(it.get('sku') or (p or {}).get('sku') or '').strip()
                full = old_name.get(oldsku)
                if full:
                    m = match_by_name(full, sheet)
                    if m:
                        name = full
            sku = m['sku'] if m else ''
            if not sku:
                if name in override:
                    # คนตัดสินใจไว้แล้วว่าชื่อนี้คือ SKU ไหน
                    sku = override[name]
                    if sku != 'NEW' and sku not in have:
                        problems.append('map.json ชี้ %r ไปที่ %s ซึ่งไม่มีในชีท' % (name[:40], sku))
                        continue
                    if sku == 'NEW':
                        sku = new_sku(name, sheet, have)
                        new_products.append({'sku': sku, 'group': 'TOOLING', 'name': name,
                                             'perPack': 1, 'unit': 'ชิ้น', 'cost': '',
                                             'price': num((p or {}).get('price')), 'reorder': 10})
                        have.add(sku)
                else:
                    # ห้ามเดา ห้ามสร้าง SKU ใหม่เอง — ของที่มีอยู่ในชีทแล้วจะกลายเป็นสองรหัส
                    problems.append('ออเดอร์ %s บรรทัด %d: จับคู่สินค้าไม่ได้ %r '
                                    '— ใส่ใน map.json ว่าเป็น SKU ไหน (หรือ "NEW" ถ้าเป็นของใหม่จริง)'
                                    % (o.get('no'), ix, name[:50]))
                    continue
            qty = int(num(it.get('qty'), 1)) or 1
            price = it.get('price')
            price = '' if price in (None, '') else num(price)
            lines.append({'sku': sku, 'qty': qty, 'price': price})
            subtotal += num(it.get('total'), (price or 0) * qty)

        chans = [c.strip() for c in str(o.get('chan') or '').split(',') if c.strip()]
        orders.append({
            'oldNo': o.get('no'), 'date': o.get('date'),
            'channel': chans[0] if chans else '',
            'extraChannels': ', '.join(chans[1:]),
            'cust': o.get('cust') or '', 'tel': o.get('tel') or '',
            'addr': o.get('addr') or '', 'carrier': carrier_of(o.get('carrier')),
            'track': o.get('track') or '', 'vat': bool(num(o.get('vat'))),
            'discount': num(o.get('disc')), 'ship': num(o.get('ship')),
            'status': status_of(o), 'pay': o.get('pay') or '',
            'memo': o.get('memo') or '', 'staff': 'ย้ายระบบ',
            'subtotal': round(subtotal, 2), 'items': lines,
        })

    channels = sorted({o['channel'] for o in orders if o['channel']})

    if problems:
        print('!! ติดปัญหา %d จุด — ยังเขียนไฟล์ให้ แต่ต้องตัดสินใจก่อนนำเข้า:' % len(problems))
        for p in problems:
            print('   -', p)

    body = [
        '/* ข้อมูลนำเข้าครั้งเดียว — สร้างจากไฟล์สำรองของแอปเดิม',
        ' *',
        ' * ไฟล์นี้มีชื่อ เบอร์ และที่อยู่ลูกค้าจริง',
        ' * ห้ามอัปขึ้น GitHub หรือที่สาธารณะใด ๆ  ใช้เสร็จให้ลบออกจากโปรเจกต์',
        ' *',
        ' * รหัสสินค้าในนี้เป็นรหัสของชีท ไม่ใช่รหัสของแอปเดิม',
        ' * (เคมี 5 ตัวรหัสไม่ตรงกัน จับคู่ด้วยชนิด+ขนาดบรรจุแทน)',
        ' */',
        'var IMPORT = {',
        '  channels: %s,' % gs(channels),
        '  products: [',
    ]
    for p in new_products:
        body.append('    %s,' % gs(p))
    body += ['  ],', '  prices: [']
    for p in prices:
        body.append('    %s,' % gs(p))
    body += ['  ],', '  lots: [']
    for l in lots:
        body.append('    %s,' % gs(l))
    body += ['  ],', '  recv: [']
    for r in recv:
        body.append('    %s,' % gs(r))
    body += ['  ],', '  orders: [']
    for o in orders:
        body.append('    %s,' % gs(o))
    body += ['  ]', '};', '']

    out.write_text('\n'.join(body), encoding='utf-8')
    print('เขียน %s' % out)
    print('  ช่องทางขายที่ต้องเพิ่ม : %s' % (', '.join(channels) or '-'))
    print('  สินค้าใหม่             : %d' % len(new_products))
    print('  ล็อต                   : %d' % len(lots))
    print('  รับเข้า                : %d' % len(recv))
    print('  ออเดอร์ (วันที่ %s)  : %d ใบ' % (today, len(orders)))
    for o in orders:
        print('    %s  %d บรรทัด  ยอดสินค้า %.2f' % (o['oldNo'], len(o['items']), o['subtotal']))
    return 1 if problems else 0


if __name__ == '__main__':
    sys.exit(main())
